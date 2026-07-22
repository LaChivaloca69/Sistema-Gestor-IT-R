"""Motor de generacion de documentos para ordenes de compra.

Convierte una plantilla (Word, Excel o PDF) y un diccionario de valores en
un PDF final listo para guardarse en ``Presupuesto.archivo_pdf``.

Convencion de campos:
    - Word (.docx) y Excel (.xlsx): marcadores ``{{nombre_campo}}`` escritos
      dentro del texto/celdas.
    - PDF: campos de formulario (AcroForm) ya definidos en el archivo.
"""
import io
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

CAMPO_PATTERN = re.compile(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\}\}")

_SOFFICE_PATHS_COMUNES = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/usr/bin/soffice",
    "/usr/lib/libreoffice/program/soffice",
    "/opt/libreoffice/program/soffice",
]


class DocumentEngineError(Exception):
    """Error controlado durante la deteccion o generacion de documentos."""


def detectar_campos(archivo, tipo_archivo):
    """Devuelve la lista de nombres de campo detectados en una plantilla.

    ``archivo`` es un objeto tipo archivo (por ejemplo un ``UploadedFile``)
    listo para leerse.
    """
    if tipo_archivo == "DOCX":
        return _detectar_campos_docx(archivo)
    if tipo_archivo == "XLSX":
        return _detectar_campos_xlsx(archivo)
    if tipo_archivo == "PDF":
        return _detectar_campos_pdf(archivo)
    raise DocumentEngineError(f"Tipo de plantilla no soportado: {tipo_archivo}")


def _detectar_campos_docx(archivo):
    from docxtpl import DocxTemplate

    archivo.seek(0)
    try:
        documento = DocxTemplate(archivo)
        campos = sorted(documento.get_undeclared_template_variables())
    except DocumentEngineError:
        raise
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo leer el archivo Word. Verifica que sea un .docx valido."
        ) from exc

    if not campos:
        raise DocumentEngineError(
            "No se encontraron campos en la plantilla. Agrega marcadores como "
            "{{folio}} o {{cliente}} dentro del documento."
        )
    return campos


def _detectar_campos_xlsx(archivo):
    import openpyxl

    archivo.seek(0)
    try:
        workbook = openpyxl.load_workbook(archivo)
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo leer el archivo Excel. Verifica que sea un .xlsx valido."
        ) from exc

    campos = set()
    for hoja in workbook.worksheets:
        for fila in hoja.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str):
                    campos.update(CAMPO_PATTERN.findall(celda.value))

    if not campos:
        raise DocumentEngineError(
            "No se encontraron campos en la plantilla. Agrega marcadores como "
            "{{folio}} o {{cliente}} dentro de las celdas."
        )
    return sorted(campos)


def _detectar_campos_pdf(archivo):
    from pypdf import PdfReader

    archivo.seek(0)
    try:
        reader = PdfReader(archivo)
        campos = reader.get_fields()
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo leer el archivo PDF. Verifica que no este danado o protegido."
        ) from exc

    if not campos:
        raise DocumentEngineError(
            "Este PDF no tiene campos de formulario rellenables. Usa una "
            "plantilla Word/Excel, o un PDF con campos de formulario (AcroForm)."
        )
    return sorted(campos.keys())


def generar_pdf(plantilla, valores):
    """Genera el PDF final a partir de una ``PlantillaDocumento`` y sus valores.

    Devuelve los bytes del PDF resultante.
    """
    tipo = plantilla.tipo_archivo
    if tipo == "PDF":
        return _generar_pdf_desde_pdf(plantilla, valores)
    if tipo == "DOCX":
        return _generar_pdf_desde_docx(plantilla, valores)
    if tipo == "XLSX":
        return _generar_pdf_desde_xlsx(plantilla, valores)
    raise DocumentEngineError(f"Tipo de plantilla no soportado: {tipo}")


def _generar_pdf_desde_pdf(plantilla, valores):
    from pypdf import PdfReader, PdfWriter

    plantilla.archivo.open("rb")
    try:
        reader = PdfReader(plantilla.archivo)
        writer = PdfWriter()
        writer.append(reader)
        writer.set_need_appearances_writer(True)

        valores_texto = {clave: "" if valor is None else str(valor) for clave, valor in valores.items()}
        for pagina in writer.pages:
            writer.update_page_form_field_values(pagina, valores_texto, auto_regenerate=False)

        buffer = io.BytesIO()
        writer.write(buffer)
        return buffer.getvalue()
    except DocumentEngineError:
        raise
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo generar el PDF a partir de la plantilla."
        ) from exc
    finally:
        plantilla.archivo.close()


def _generar_pdf_desde_docx(plantilla, valores):
    from docxtpl import DocxTemplate

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        plantilla.archivo.open("rb")
        try:
            documento = DocxTemplate(plantilla.archivo)
            documento.render(valores)
        except Exception as exc:
            raise DocumentEngineError(
                "No se pudo rellenar la plantilla Word con los datos capturados."
            ) from exc
        finally:
            plantilla.archivo.close()

        docx_path = tmp_dir_path / "orden_compra.docx"
        documento.save(docx_path)
        return _convertir_a_pdf_con_libreoffice(docx_path, tmp_dir_path)


def _generar_pdf_desde_xlsx(plantilla, valores):
    import openpyxl

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        plantilla.archivo.open("rb")
        try:
            workbook = openpyxl.load_workbook(plantilla.archivo)
        except Exception as exc:
            raise DocumentEngineError(
                "No se pudo leer la plantilla Excel para rellenarla."
            ) from exc
        finally:
            plantilla.archivo.close()

        def reemplazar(match):
            return str(valores.get(match.group(1), ""))

        for hoja in workbook.worksheets:
            for fila in hoja.iter_rows():
                for celda in fila:
                    if isinstance(celda.value, str) and "{{" in celda.value:
                        celda.value = CAMPO_PATTERN.sub(reemplazar, celda.value)

        xlsx_path = tmp_dir_path / "orden_compra.xlsx"
        workbook.save(xlsx_path)
        return _convertir_a_pdf_con_libreoffice(xlsx_path, tmp_dir_path)


def _encontrar_soffice():
    encontrado = shutil.which("soffice") or shutil.which("soffice.exe")
    if encontrado:
        return encontrado
    for ruta in _SOFFICE_PATHS_COMUNES:
        if Path(ruta).exists():
            return ruta
    return None


def _convertir_a_pdf_con_libreoffice(archivo_origen, tmp_dir_path):
    soffice = _encontrar_soffice()
    if not soffice:
        raise DocumentEngineError(
            "No se encontro LibreOffice instalado en el servidor. Instalalo "
            "(por ejemplo 'apt-get install libreoffice' en Linux, o "
            "'winget install TheDocumentFoundation.LibreOffice' en Windows) "
            "para poder generar ordenes de compra desde plantillas Word o Excel."
        )

    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                "--norestore",
                "--convert-to", "pdf",
                "--outdir", str(tmp_dir_path),
                str(archivo_origen),
            ],
            check=True,
            capture_output=True,
            timeout=120,
        )
    except subprocess.CalledProcessError as exc:
        raise DocumentEngineError(
            "No se pudo convertir el documento a PDF. Intenta de nuevo o "
            "revisa que la plantilla no este danada."
        ) from exc
    except subprocess.TimeoutExpired as exc:
        raise DocumentEngineError(
            "La conversion a PDF esta tardando demasiado. Intenta de nuevo."
        ) from exc

    pdf_path = archivo_origen.with_suffix(".pdf")
    if not pdf_path.exists():
        raise DocumentEngineError("No se genero el PDF esperado durante la conversion.")

    return pdf_path.read_bytes()
