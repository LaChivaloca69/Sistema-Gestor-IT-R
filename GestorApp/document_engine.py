"""Motor de generacion de documentos para ordenes de compra.

Convierte una plantilla (Word, Excel o PDF) y un diccionario de valores en
un PDF final listo para guardarse en ``OrdenCompra.archivo_pdf``.

Convencion de campos:
    - Word (.docx) y Excel (.xlsx): marcadores ``{{nombre_campo}}`` escritos
      dentro del texto/celdas.
    - PDF: campos de formulario (AcroForm) ya definidos en el archivo.
    - Para ordenes creadas desde el modelo se usan marcadores fijos
      (``{{folio}}``, ``{{id_1}}`` ... ``{{amount_15}}``, etc.).
"""
import io
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

CAMPO_PATTERN = re.compile(r"\{\{\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*\}\}")
MAX_LINEAS_PLANTILLA = 20

_SOFFICE_PATHS_COMUNES = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/usr/bin/soffice",
    "/usr/lib/libreoffice/program/soffice",
    "/opt/libreoffice/program/soffice",
]

DEFAULT_TEMPLATE_RELATIVE = Path("plantillas_orden_compra") / "Formato_Orden_Compra.xlsx"


class DocumentEngineError(Exception):
    """Error controlado durante la deteccion o generacion de documentos."""


def detectar_campos(archivo, tipo_archivo):
    """Devuelve la lista de nombres de campo detectados en una plantilla.

    ``archivo`` es un objeto tipo archivo (por ejemplo un ``UploadedFile``)
    listo para leerse.
    """
    if tipo_archivo == "DOCX":
        return _detectar_campos_docx(archivo)
    if tipo_archivo == "XLSX":
        return _detectar_campos_xlsx(archivo)
    if tipo_archivo == "PDF":
        return _detectar_campos_pdf(archivo)
    raise DocumentEngineError(f"Tipo de plantilla no soportado: {tipo_archivo}")


def _detectar_campos_docx(archivo):
    from docxtpl import DocxTemplate

    archivo.seek(0)
    try:
        documento = DocxTemplate(archivo)
        campos = sorted(documento.get_undeclared_template_variables())
    except DocumentEngineError:
        raise
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo leer el archivo Word. Verifica que sea un .docx valido."
        ) from exc

    if not campos:
        raise DocumentEngineError(
            "No se encontraron campos en la plantilla. Agrega marcadores como "
            "{{folio}} o {{cliente}} dentro del documento."
        )
    return campos


def _detectar_campos_xlsx(archivo):
    import openpyxl

    archivo.seek(0)
    try:
        workbook = openpyxl.load_workbook(archivo)
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo leer el archivo Excel. Verifica que sea un .xlsx valido."
        ) from exc

    campos = set()
    for hoja in workbook.worksheets:
        for fila in hoja.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str):
                    campos.update(CAMPO_PATTERN.findall(celda.value))

    if not campos:
        raise DocumentEngineError(
            "No se encontraron campos en la plantilla. Agrega marcadores como "
            "{{folio}} o {{cliente}} dentro de las celdas."
        )
    return sorted(campos)


def _detectar_campos_pdf(archivo):
    from pypdf import PdfReader

    archivo.seek(0)
    try:
        reader = PdfReader(archivo)
        campos = reader.get_fields()
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo leer el archivo PDF. Verifica que no este danado o protegido."
        ) from exc

    if not campos:
        raise DocumentEngineError(
            "Este PDF no tiene campos de formulario rellenables. Usa una "
            "plantilla Word/Excel, o un PDF con campos de formulario (AcroForm)."
        )
    return sorted(campos.keys())


def generar_pdf(plantilla, valores):
    """Genera el PDF final a partir de una ``PlantillaDocumento`` y sus valores.

    Devuelve los bytes del PDF resultante.
    """
    tipo = plantilla.tipo_archivo
    if tipo == "PDF":
        return _generar_pdf_desde_pdf(plantilla, valores)
    if tipo == "DOCX":
        return _generar_pdf_desde_docx(plantilla, valores)
    if tipo == "XLSX":
        return _generar_pdf_desde_xlsx(plantilla, valores)
    raise DocumentEngineError(f"Tipo de plantilla no soportado: {tipo}")


def _generar_pdf_desde_pdf(plantilla, valores):
    from pypdf import PdfReader, PdfWriter

    plantilla.archivo.open("rb")
    try:
        reader = PdfReader(plantilla.archivo)
        writer = PdfWriter()
        writer.append(reader)
        writer.set_need_appearances_writer(True)

        valores_texto = {clave: "" if valor is None else str(valor) for clave, valor in valores.items()}
        for pagina in writer.pages:
            writer.update_page_form_field_values(pagina, valores_texto, auto_regenerate=False)

        buffer = io.BytesIO()
        writer.write(buffer)
        return buffer.getvalue()
    except DocumentEngineError:
        raise
    except Exception as exc:
        raise DocumentEngineError(
            "No se pudo generar el PDF a partir de la plantilla."
        ) from exc
    finally:
        plantilla.archivo.close()


def _generar_pdf_desde_docx(plantilla, valores):
    from docxtpl import DocxTemplate

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        plantilla.archivo.open("rb")
        try:
            documento = DocxTemplate(plantilla.archivo)
            documento.render(valores)
        except Exception as exc:
            raise DocumentEngineError(
                "No se pudo rellenar la plantilla Word con los datos capturados."
            ) from exc
        finally:
            plantilla.archivo.close()

        docx_path = tmp_dir_path / "orden_compra.docx"
        documento.save(docx_path)
        return _convertir_a_pdf_con_libreoffice(docx_path, tmp_dir_path)


def _rellenar_xlsx_preservando(origen, destino, valores):
    """Copia el .xlsx y reemplaza ``{{marcadores}}`` en XML sin perder imagenes.

    Evita reescribir el libro con openpyxl (que puede descartar dibujos/logos).
    """
    import zipfile

    def reemplazar(match):
        return str(valores.get(match.group(1), ""))

    with zipfile.ZipFile(origen, "r") as zin, zipfile.ZipFile(destino, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            nombre = item.filename.lower()
            if nombre.endswith(".xml") or nombre.endswith(".rels"):
                try:
                    texto = data.decode("utf-8")
                except UnicodeDecodeError:
                    zout.writestr(item, data)
                    continue
                if "{{" in texto:
                    texto = CAMPO_PATTERN.sub(reemplazar, texto)
                    data = texto.encode("utf-8")
            zout.writestr(item, data)


def _generar_pdf_desde_xlsx(plantilla, valores):
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        xlsx_path = tmp_dir_path / "orden_compra.xlsx"
        origen_tmp = tmp_dir_path / "origen.xlsx"

        plantilla.archivo.open("rb")
        try:
            origen_tmp.write_bytes(plantilla.archivo.read())
        except Exception as exc:
            raise DocumentEngineError(
                "No se pudo leer la plantilla Excel para rellenarla."
            ) from exc
        finally:
            plantilla.archivo.close()

        try:
            _rellenar_xlsx_preservando(origen_tmp, xlsx_path, valores)
        except Exception as exc:
            raise DocumentEngineError(
                "No se pudo rellenar la plantilla Excel conservando su formato."
            ) from exc

        return _convertir_a_pdf_con_libreoffice(xlsx_path, tmp_dir_path)


def _encontrar_soffice():
    encontrado = shutil.which("soffice") or shutil.which("soffice.exe")
    if encontrado:
        return encontrado
    for ruta in _SOFFICE_PATHS_COMUNES:
        if Path(ruta).exists():
            return ruta
    return None


def _convertir_a_pdf_con_libreoffice(archivo_origen, tmp_dir_path):
    soffice = _encontrar_soffice()
    if not soffice:
        raise DocumentEngineError(
            "No se encontro LibreOffice instalado en el servidor. Instalalo "
            "(por ejemplo 'apt-get install libreoffice' en Linux, o "
            "'winget install TheDocumentFoundation.LibreOffice' en Windows) "
            "para poder generar ordenes de compra desde plantillas Word o Excel."
        )

    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                "--norestore",
                "--convert-to", "pdf",
                "--outdir", str(tmp_dir_path),
                str(archivo_origen),
            ],
            check=True,
            capture_output=True,
            timeout=120,
        )
    except subprocess.CalledProcessError as exc:
        raise DocumentEngineError(
            "No se pudo convertir el documento a PDF. Intenta de nuevo o "
            "revisa que la plantilla no este danada."
        ) from exc
    except subprocess.TimeoutExpired as exc:
        raise DocumentEngineError(
            "La conversion a PDF esta tardando demasiado. Intenta de nuevo."
        ) from exc

    pdf_path = archivo_origen.with_suffix(".pdf")
    if not pdf_path.exists():
        raise DocumentEngineError("No se genero el PDF esperado durante la conversion.")

    return pdf_path.read_bytes()


def valores_desde_orden(orden):
    """Construye el diccionario de marcadores a partir de una OrdenCompra."""
    proveedor = orden.proveedor
    elaborado = ""
    if orden.elaborado_por_id:
        elaborado = orden.elaborado_por.get_full_name() or orden.elaborado_por.get_username()

    valores = {
        "folio": orden.folio_orden or "",
        "fecha": orden.fecha.strftime("%Y-%m-%d") if orden.fecha else "",
        "elaborado_por": elaborado,
        "proveedor_nombre": proveedor.nombre_proveedor if proveedor else "",
        "proveedor_contacto": (proveedor.contacto or "") if proveedor else "",
        "proveedor_telefono": (proveedor.telefono or "") if proveedor else "",
        "proveedor_email": (proveedor.correo or "") if proveedor else "",
        "comentarios": orden.comentarios or "",
        "notas": orden.notas or "",
        "moneda": orden.tipo_moneda or "",
        "iva_porcentaje": str(orden.iva_porcentaje or ""),
        "subtotal": f"{orden.subtotal:.2f}",
        "iva": f"{orden.iva_monto:.2f}",
        "total": f"{orden.total:.2f}",
    }

    for i in range(1, MAX_LINEAS_PLANTILLA + 1):
        valores[f"id_{i}"] = ""
        valores[f"descripcion_{i}"] = ""
        valores[f"cantidad_{i}"] = ""
        valores[f"pu_{i}"] = ""
        valores[f"amount_{i}"] = ""

    for idx, detalle in enumerate(orden.detalles.all()[:MAX_LINEAS_PLANTILLA], start=1):
        valores[f"id_{idx}"] = detalle.id_producto or ""
        valores[f"descripcion_{idx}"] = detalle.descripcion or ""
        valores[f"cantidad_{idx}"] = f"{detalle.cantidad:.2f}"
        valores[f"pu_{idx}"] = f"{detalle.precio_unitario:.2f}"
        valores[f"amount_{idx}"] = f"{detalle.importe:.2f}"

    return valores


def _ruta_plantilla_default():
    from django.conf import settings

    return Path(settings.MEDIA_ROOT) / DEFAULT_TEMPLATE_RELATIVE


def asegurar_plantilla_default():
    """Crea el Excel base con marcadores si aun no existe."""
    ruta = _ruta_plantilla_default()
    if ruta.exists():
        return ruta

    import openpyxl
    from openpyxl.styles import Alignment, Font

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orden de compra"

    ws["A1"] = "ORDEN DE COMPRA"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Folio:"
    ws["B3"] = "{{folio}}"
    ws["D3"] = "Fecha:"
    ws["E3"] = "{{fecha}}"
    ws["A4"] = "Elaborado por:"
    ws["B4"] = "{{elaborado_por}}"
    ws["D4"] = "Moneda:"
    ws["E4"] = "{{moneda}}"

    ws["A6"] = "Proveedor:"
    ws["B6"] = "{{proveedor_nombre}}"
    ws["A7"] = "Contacto:"
    ws["B7"] = "{{proveedor_contacto}}"
    ws["A8"] = "Telefono:"
    ws["B8"] = "{{proveedor_telefono}}"
    ws["A9"] = "Email:"
    ws["B9"] = "{{proveedor_email}}"

    headers = ["ID", "Descripcion", "Cantidad", "P.U.", "Amount"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = Font(bold=True)

    for i in range(1, MAX_LINEAS_PLANTILLA + 1):
        row = 11 + i
        ws.cell(row=row, column=1, value=f"{{{{id_{i}}}}}")
        ws.cell(row=row, column=2, value=f"{{{{descripcion_{i}}}}}")
        ws.cell(row=row, column=3, value=f"{{{{cantidad_{i}}}}}")
        ws.cell(row=row, column=4, value=f"{{{{pu_{i}}}}}")
        ws.cell(row=row, column=5, value=f"{{{{amount_{i}}}}}")

    base = 12 + MAX_LINEAS_PLANTILLA
    ws.cell(row=base, column=4, value="Subtotal")
    ws.cell(row=base, column=5, value="{{subtotal}}")
    ws.cell(row=base + 1, column=4, value="IVA {{iva_porcentaje}}%")
    ws.cell(row=base + 1, column=5, value="{{iva}}")
    ws.cell(row=base + 2, column=4, value="Total")
    ws.cell(row=base + 2, column=5, value="{{total}}")
    ws.cell(row=base + 4, column=1, value="Comentarios:")
    ws.cell(row=base + 5, column=1, value="{{comentarios}}")

    for col, width in enumerate([16, 40, 12, 14, 14], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(ruta)
    return ruta


def valores_desde_payload(payload):
    """Construye marcadores a partir de un dict (preview AJAX, sin modelo guardado)."""
    from decimal import Decimal

    def _txt(key, default=""):
        val = payload.get(key, default)
        return "" if val is None else str(val)

    detalles = payload.get("detalles") or []
    subtotal = Decimal("0")
    lineas = []
    for item in detalles[:MAX_LINEAS_PLANTILLA]:
        try:
            cantidad = Decimal(str(item.get("cantidad") or 0))
        except Exception:
            cantidad = Decimal("0")
        try:
            pu = Decimal(str(item.get("precio_unitario") or 0))
        except Exception:
            pu = Decimal("0")
        importe = (cantidad * pu).quantize(Decimal("0.01"))
        subtotal += importe
        lineas.append(
            {
                "id_producto": item.get("id_producto") or "",
                "descripcion": item.get("descripcion") or "",
                "cantidad": cantidad,
                "precio_unitario": pu,
                "importe": importe,
            }
        )

    try:
        iva_porcentaje = Decimal(str(payload.get("iva_porcentaje") or 0))
    except Exception:
        iva_porcentaje = Decimal("0")
    iva_monto = (subtotal * iva_porcentaje / Decimal("100")).quantize(Decimal("0.01"))
    total = subtotal + iva_monto

    valores = {
        "folio": _txt("folio") or "(auto)",
        "fecha": _txt("fecha"),
        "elaborado_por": _txt("elaborado_por"),
        "proveedor_nombre": _txt("proveedor_nombre"),
        "proveedor_contacto": _txt("proveedor_contacto"),
        "proveedor_telefono": _txt("proveedor_telefono"),
        "proveedor_email": _txt("proveedor_email"),
        "comentarios": _txt("comentarios"),
        "notas": _txt("notas"),
        "moneda": _txt("moneda", "MXN"),
        "iva_porcentaje": f"{iva_porcentaje:.2f}".rstrip("0").rstrip("."),
        "subtotal": f"{subtotal:.2f}",
        "iva": f"{iva_monto:.2f}",
        "total": f"{total:.2f}",
    }

    for i in range(1, MAX_LINEAS_PLANTILLA + 1):
        valores[f"id_{i}"] = ""
        valores[f"descripcion_{i}"] = ""
        valores[f"cantidad_{i}"] = ""
        valores[f"pu_{i}"] = ""
        valores[f"amount_{i}"] = ""

    for idx, linea in enumerate(lineas, start=1):
        valores[f"id_{idx}"] = linea["id_producto"]
        valores[f"descripcion_{idx}"] = linea["descripcion"]
        valores[f"cantidad_{idx}"] = f"{linea['cantidad']:.2f}"
        valores[f"pu_{idx}"] = f"{linea['precio_unitario']:.2f}"
        valores[f"amount_{idx}"] = f"{linea['importe']:.2f}"

    return valores


def _aplicar_valores_a_workbook(workbook, valores):
    def reemplazar(match):
        return str(valores.get(match.group(1), ""))

    for hoja in workbook.worksheets:
        for fila in hoja.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str) and "{{" in celda.value:
                    celda.value = CAMPO_PATTERN.sub(reemplazar, celda.value)


def _celda_a_texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, float):
        return f"{valor:.2f}"
    return str(valor)


def workbook_a_html(workbook, max_row=52, max_col=13):
    """Convierte la hoja principal de un workbook a HTML (tabla)."""
    from html import escape
    from openpyxl.utils import range_boundaries

    if "FORMATO" in workbook.sheetnames:
        ws = workbook["FORMATO"]
    else:
        ws = workbook.active

    merge_anchor = {}
    skip = set()
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_c, max_r = range_boundaries(str(merged))
        if min_row > max_row or min_col > max_col:
            continue
        rowspan = min(max_r, max_row) - min_row + 1
        colspan = min(max_c, max_col) - min_col + 1
        merge_anchor[(min_row, min_col)] = (rowspan, colspan)
        for r in range(min_row, min(max_r, max_row) + 1):
            for c in range(min_col, min(max_c, max_col) + 1):
                if (r, c) != (min_row, min_col):
                    skip.add((r, c))

    parts = [
        '<div class="oc-xlsx-preview">',
        f'<table class="oc-xlsx-table" data-sheet="{escape(ws.title)}">',
    ]
    for r in range(1, max_row + 1):
        parts.append("<tr>")
        for c in range(1, max_col + 1):
            if (r, c) in skip:
                continue
            cell = ws.cell(row=r, column=c)
            texto = escape(_celda_a_texto(cell.value)).replace("\n", "<br>")
            attrs = []
            if (r, c) in merge_anchor:
                rowspan, colspan = merge_anchor[(r, c)]
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')
            attr_str = (" " + " ".join(attrs)) if attrs else ""
            parts.append(f"<td{attr_str}>{texto}</td>")
        parts.append("</tr>")
    parts.append("</table></div>")
    return "".join(parts)


def render_preview_pdf(valores, plantilla=None):
    """Rellena la plantilla (con imagenes/estilos) y la convierte a PDF para vista previa."""
    if plantilla is not None and plantilla.archivo:
        return generar_pdf(plantilla, valores)

    ruta = asegurar_plantilla_default()
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        xlsx_path = tmp_dir_path / "orden_compra_preview.xlsx"
        try:
            _rellenar_xlsx_preservando(ruta, xlsx_path, valores)
        except Exception as exc:
            raise DocumentEngineError(
                "No se pudo preparar la plantilla por defecto para la vista previa."
            ) from exc
        return _convertir_a_pdf_con_libreoffice(xlsx_path, tmp_dir_path)


def generar_pdf_orden_compra(orden):
    """Genera el PDF de una OrdenCompra usando su plantilla o la default."""
    valores = valores_desde_orden(orden)
    return render_preview_pdf(valores, plantilla=orden.plantilla if orden.plantilla_id else None)
