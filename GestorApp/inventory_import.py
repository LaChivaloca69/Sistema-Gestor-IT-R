"""Importacion masiva de equipos y perifericos desde Excel."""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import historial
from .models import (
    AsignacionEquipo,
    CategoriaEquipo,
    Equipo,
    EstadoAsignacion,
    EstadoEquipo,
    ModuloHistorial,
    OrigenAltaEquipo,
    Personal,
    Proveedor,
    TipoCategoriaInventario,
    TipoMovimiento,
    Ubicacion,
)

MAX_IMPORT_ROWS = 2000
MAX_FILE_BYTES = 5 * 1024 * 1024

SHEET_EQUIPOS = "Equipos"
SHEET_PERIFERICOS = "Perifericos"
SHEET_CATALOGOS = "Catalogos"

EQUIPOS_HEADERS = [
    "codigo_inventario",
    "categoria",
    "marca",
    "modelo",
    "numero_serie",
    "estado",
    "ubicacion",
    "numero_empleado",
    "proveedor",
    "descripcion",
    "origen_alta",
]

PERIFERICOS_HEADERS = [
    "codigo_inventario",
    "categoria",
    "marca",
    "modelo",
    "numero_serie",
    "estado",
    "ubicacion",
    "descripcion",
]

STATUS_OK = "ok"
STATUS_WARNING = "warning"
STATUS_ERROR = "error"


def _norm_key(value: str) -> str:
    text = (value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass
class ImportRow:
    sheet: str
    row_num: int
    tipo: str
    codigo: str = ""
    categoria: str = ""
    status: str = STATUS_OK
    messages: list[str] = field(default_factory=list)
    payload: dict[str, Any] | None = None

    def to_session(self) -> dict:
        return {
            "sheet": self.sheet,
            "row_num": self.row_num,
            "tipo": self.tipo,
            "codigo": self.codigo,
            "categoria": self.categoria,
            "status": self.status,
            "messages": self.messages,
            "payload": self.payload,
        }

    @classmethod
    def from_session(cls, data: dict) -> ImportRow:
        return cls(
            sheet=data.get("sheet", ""),
            row_num=int(data.get("row_num") or 0),
            tipo=data.get("tipo", ""),
            codigo=data.get("codigo", ""),
            categoria=data.get("categoria", ""),
            status=data.get("status", STATUS_ERROR),
            messages=list(data.get("messages") or []),
            payload=data.get("payload"),
        )


@dataclass
class ImportCatalogs:
    categorias_equipo: dict[str, CategoriaEquipo]
    categorias_periferico: dict[str, CategoriaEquipo]
    ubicaciones: dict[str, Ubicacion]
    personal: dict[str, Personal]
    proveedores: dict[str, Proveedor]
    codigos_db: set[str]
    series_db: set[str]

    @classmethod
    def load(cls) -> ImportCatalogs:
        categorias_equipo = {}
        categorias_periferico = {}
        for cat in CategoriaEquipo.objects.filter(activo=True):
            key = _norm_key(cat.nombre_categoria)
            if cat.tipo == TipoCategoriaInventario.EQUIPO:
                categorias_equipo[key] = cat
            elif cat.tipo == TipoCategoriaInventario.PERIFERICO:
                categorias_periferico[key] = cat

        ubicaciones = {}
        for ub in Ubicacion.objects.filter(activo=True).select_related(
            "edificio", "zona"
        ):
            ubicaciones[_norm_key(str(ub))] = ub
            if ub.referencia:
                ubicaciones[_norm_key(ub.referencia)] = ub

        personal = {
            p.numero_empleado.strip().lower(): p
            for p in Personal.objects.filter(activo=True)
            if p.numero_empleado
        }

        proveedores = {}
        for prov in Proveedor.objects.filter(activo=True):
            proveedores[_norm_key(prov.nombre_proveedor)] = prov
            if prov.codigo_interno:
                proveedores[_norm_key(prov.codigo_interno)] = prov
            proveedores[_norm_key(str(prov))] = prov

        codigos_db = set(
            Equipo.objects.values_list("codigo_inventario", flat=True)
        )
        series_db = {
            s.lower()
            for s in Equipo.objects.exclude(numero_serie__isnull=True)
            .exclude(numero_serie="")
            .values_list("numero_serie", flat=True)
        }

        return cls(
            categorias_equipo=categorias_equipo,
            categorias_periferico=categorias_periferico,
            ubicaciones=ubicaciones,
            personal=personal,
            proveedores=proveedores,
            codigos_db=codigos_db,
            series_db=series_db,
        )


def _estado_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for value, label in EstadoEquipo.choices:
        mapping[_norm_key(value)] = value
        mapping[_norm_key(label)] = value
    return mapping


def _origen_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for value, label in OrigenAltaEquipo.choices:
        mapping[_norm_key(value)] = value
        mapping[_norm_key(label)] = value
        if "/" in label:
            mapping[_norm_key(label.split("/")[0].strip())] = value
    return mapping


def _read_sheet_rows(workbook, sheet_name: str, expected_headers: list[str]):
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_map = {}
    for idx, cell in enumerate(rows[0]):
        key = _norm_key(_cell_str(cell))
        if key:
            header_map[key] = idx

    parsed = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(_cell_str(cell) == "" for cell in row):
            continue
        data = {}
        for header in expected_headers:
            col = header_map.get(_norm_key(header))
            data[header] = _cell_str(row[col]) if col is not None and col < len(row) else ""
        parsed.append((row_idx, data))
    return parsed


def _add_message(row: ImportRow, level: str, text: str):
    row.messages.append(text)
    if level == STATUS_ERROR:
        row.status = STATUS_ERROR
    elif level == STATUS_WARNING and row.status != STATUS_ERROR:
        row.status = STATUS_WARNING


def _validate_row(
    *,
    sheet: str,
    row_num: int,
    tipo: str,
    data: dict[str, str],
    catalogs: ImportCatalogs,
    codigos_archivo: set[str],
    series_archivo: set[str],
) -> ImportRow:
    import_row = ImportRow(
        sheet=sheet,
        row_num=row_num,
        tipo=tipo,
        codigo=data.get("codigo_inventario", ""),
        categoria=data.get("categoria", ""),
    )

    codigo = data.get("codigo_inventario", "").strip()
    if not codigo:
        _add_message(import_row, STATUS_ERROR, "Falta codigo_inventario.")
        return import_row

    codigo_key = codigo.lower()
    if codigo_key in codigos_archivo:
        _add_message(import_row, STATUS_ERROR, "Codigo duplicado en el archivo.")
    if codigo in catalogs.codigos_db:
        _add_message(import_row, STATUS_ERROR, "El codigo ya existe en el sistema.")

    categoria_nombre = data.get("categoria", "").strip()
    if not categoria_nombre:
        _add_message(import_row, STATUS_ERROR, "Falta categoria.")
        categoria = None
    else:
        cat_map = (
            catalogs.categorias_equipo
            if tipo == TipoCategoriaInventario.EQUIPO
            else catalogs.categorias_periferico
        )
        categoria = cat_map.get(_norm_key(categoria_nombre))
        if categoria is None:
            _add_message(
                import_row,
                STATUS_ERROR,
                f"Categoria '{categoria_nombre}' no existe para tipo {tipo}.",
            )

    serie = data.get("numero_serie", "").strip()
    if serie:
        serie_key = serie.lower()
        if serie_key in series_archivo:
            _add_message(import_row, STATUS_ERROR, "Numero de serie duplicado en el archivo.")
        if serie_key in catalogs.series_db:
            _add_message(import_row, STATUS_ERROR, "El numero de serie ya existe en el sistema.")

    estado_raw = data.get("estado", "").strip()
    estado = EstadoEquipo.DISPONIBLE
    if estado_raw:
        estado = _estado_map().get(_norm_key(estado_raw))
        if estado is None:
            _add_message(import_row, STATUS_ERROR, f"Estado invalido: {estado_raw}.")
        elif estado in {EstadoEquipo.BAJA, EstadoEquipo.EN_MANTENIMIENTO}:
            _add_message(
                import_row,
                STATUS_ERROR,
                "No se importan equipos en Baja o En Mantenimiento.",
            )

    ubicacion = None
    ubicacion_raw = data.get("ubicacion", "").strip()
    if ubicacion_raw:
        ubicacion = catalogs.ubicaciones.get(_norm_key(ubicacion_raw))
        if ubicacion is None:
            _add_message(
                import_row,
                STATUS_ERROR,
                f"Ubicacion no encontrada: {ubicacion_raw}.",
            )

    personal = None
    numero_empleado = data.get("numero_empleado", "").strip()
    if numero_empleado:
        if tipo == TipoCategoriaInventario.PERIFERICO:
            _add_message(
                import_row,
                STATUS_WARNING,
                "numero_empleado ignorado en perifericos; vincule al kit manualmente.",
            )
        else:
            personal = catalogs.personal.get(numero_empleado.lower())
            if personal is None:
                _add_message(
                    import_row,
                    STATUS_ERROR,
                    f"Personal no encontrado: {numero_empleado}.",
                )

    if estado == EstadoEquipo.ASIGNADO and personal is None and tipo == TipoCategoriaInventario.EQUIPO:
        _add_message(
            import_row,
            STATUS_WARNING,
            "Estado Asignado sin numero_empleado; se creara En Stock.",
        )
        estado = EstadoEquipo.DISPONIBLE

    proveedor = None
    proveedor_raw = data.get("proveedor", "").strip()
    if proveedor_raw:
        proveedor = catalogs.proveedores.get(_norm_key(proveedor_raw))
        if proveedor is None:
            _add_message(
                import_row,
                STATUS_ERROR,
                f"Proveedor no encontrado: {proveedor_raw}.",
            )

    origen = OrigenAltaEquipo.LEGADO
    if tipo == TipoCategoriaInventario.EQUIPO:
        origen_raw = data.get("origen_alta", "").strip()
        if origen_raw:
            origen = _origen_map().get(_norm_key(origen_raw))
            if origen is None:
                _add_message(
                    import_row,
                    STATUS_ERROR,
                    f"Origen de alta invalido: {origen_raw}.",
                )

    if import_row.status == STATUS_ERROR:
        return import_row

    if ubicacion is None and personal is None:
        from .views.helpers import _get_espacio_stock_default

        stock = _get_espacio_stock_default()
        if stock:
            ubicacion = stock
        else:
            _add_message(
                import_row,
                STATUS_WARNING,
                "Sin ubicacion; quedara sin espacio fisico.",
            )

    import_row.payload = {
        "codigo_inventario": codigo,
        "categoria_id": categoria.pk if categoria else None,
        "marca": data.get("marca", "").strip() or None,
        "modelo": data.get("modelo", "").strip() or None,
        "numero_serie": serie or None,
        "estado_equipo": estado,
        "ubicacion_id": ubicacion.pk if ubicacion else None,
        "personal_id": personal.pk if personal else None,
        "proveedor_id": proveedor.pk if proveedor else None,
        "descripcion_equipo": data.get("descripcion", "").strip() or None,
        "origen_alta": origen,
    }
    return import_row


def parse_import_workbook(file_obj) -> tuple[list[ImportRow], list[str]]:
    """Lee y valida un workbook. Devuelve filas y errores globales."""
    global_errors: list[str] = []
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception:
        return [], ["No se pudo leer el archivo Excel. Use un .xlsx valido."]

    equipos_raw = _read_sheet_rows(workbook, SHEET_EQUIPOS, EQUIPOS_HEADERS)
    perifericos_raw = _read_sheet_rows(workbook, SHEET_PERIFERICOS, PERIFERICOS_HEADERS)
    workbook.close()

    total = len(equipos_raw) + len(perifericos_raw)
    if total == 0:
        return [], ["El archivo no tiene filas en las hojas Equipos o Perifericos."]
    if total > MAX_IMPORT_ROWS:
        return [], [f"Maximo {MAX_IMPORT_ROWS} filas por importacion (tiene {total})."]

    catalogs = ImportCatalogs.load()
    codigos_archivo: set[str] = set()
    series_archivo: set[str] = set()
    rows: list[ImportRow] = []

    for row_num, data in equipos_raw:
        row = _validate_row(
            sheet=SHEET_EQUIPOS,
            row_num=row_num,
            tipo=TipoCategoriaInventario.EQUIPO,
            data=data,
            catalogs=catalogs,
            codigos_archivo=codigos_archivo,
            series_archivo=series_archivo,
        )
        rows.append(row)
        if row.codigo:
            codigos_archivo.add(row.codigo.lower())
        serie_raw = data.get("numero_serie", "").strip()
        if serie_raw:
            series_archivo.add(serie_raw.lower())

    for row_num, data in perifericos_raw:
        row = _validate_row(
            sheet=SHEET_PERIFERICOS,
            row_num=row_num,
            tipo=TipoCategoriaInventario.PERIFERICO,
            data=data,
            catalogs=catalogs,
            codigos_archivo=codigos_archivo,
            series_archivo=series_archivo,
        )
        rows.append(row)
        if row.codigo:
            codigos_archivo.add(row.codigo.lower())
        serie_raw = data.get("numero_serie", "").strip()
        if serie_raw:
            series_archivo.add(serie_raw.lower())

    return rows, global_errors


def summarize_rows(rows: list[ImportRow]) -> dict[str, int]:
    equipos = [r for r in rows if r.tipo == TipoCategoriaInventario.EQUIPO]
    perifericos = [r for r in rows if r.tipo == TipoCategoriaInventario.PERIFERICO]
    importables = [r for r in rows if r.status in {STATUS_OK, STATUS_WARNING}]
    return {
        "total": len(rows),
        "equipos": len(equipos),
        "perifericos": len(perifericos),
        "ok": sum(1 for r in rows if r.status == STATUS_OK),
        "warnings": sum(1 for r in rows if r.status == STATUS_WARNING),
        "errors": sum(1 for r in rows if r.status == STATUS_ERROR),
        "importables": len(importables),
    }


def execute_import(rows: list[ImportRow], request=None) -> dict[str, Any]:
    """Crea registros para filas validas. Devuelve resumen de ejecucion."""
    from .views.helpers import (
        _aplicar_asignacion_a_equipo,
        _crear_movimiento,
        _get_equipo_responsable,
        _reconciliar_estado_equipo,
    )

    result = {
        "equipos_creados": 0,
        "perifericos_creados": 0,
        "asignaciones_creadas": 0,
        "errores_ejecucion": [],
        "filas_omitidas": 0,
    }

    for row in rows:
        if row.status == STATUS_ERROR or not row.payload:
            result["filas_omitidas"] += 1
            continue
        try:
            with transaction.atomic():
                payload = row.payload
                equipo = Equipo.objects.create(
                    codigo_inventario=payload["codigo_inventario"],
                    categoria_id=payload["categoria_id"],
                    marca=payload["marca"],
                    modelo=payload["modelo"],
                    numero_serie=payload["numero_serie"],
                    estado_equipo=payload["estado_equipo"],
                    ubicacion_id=payload["ubicacion_id"],
                    proveedor_id=payload["proveedor_id"],
                    descripcion_equipo=payload["descripcion_equipo"],
                    origen_alta=payload.get("origen_alta") or OrigenAltaEquipo.LEGADO,
                    fecha_alta=timezone.localdate(),
                    activo=True,
                )
                _reconciliar_estado_equipo(equipo)

                titulo_tipo = (
                    "Equipo"
                    if row.tipo == TipoCategoriaInventario.EQUIPO
                    else "Periferico"
                )
                historial.registrar_creacion(
                    request,
                    modulo=ModuloHistorial.EQUIPO,
                    titulo=f"{titulo_tipo} importado: {equipo.codigo_inventario}",
                    objeto=equipo,
                    enlace_nombre="equipo_detail",
                    metadata={
                        "origen": "importacion_excel",
                        "hoja": row.sheet,
                        "fila": row.row_num,
                        "tipo_inventario": row.tipo,
                    },
                )
                _crear_movimiento(
                    equipo,
                    TipoMovimiento.DADA_DE_ALTA,
                    origen=None,
                    destino=equipo.ubicacion,
                    responsable=_get_equipo_responsable(equipo),
                    observaciones="Importacion masiva desde Excel.",
                    request=request,
                )

                personal_id = payload.get("personal_id")
                if personal_id and row.tipo == TipoCategoriaInventario.EQUIPO:
                    personal = Personal.objects.filter(pk=personal_id).first()
                    if personal:
                        asignacion = AsignacionEquipo.objects.create(
                            equipo=equipo,
                            personal=personal,
                            estado_asignacion=EstadoAsignacion.ACTIVA,
                            observaciones="Importacion masiva desde Excel.",
                        )
                        _reconciliar_estado_equipo(equipo)
                        ubicacion_anterior, ubicacion_nueva = _aplicar_asignacion_a_equipo(
                            equipo, personal, request=request
                        )
                        _crear_movimiento(
                            equipo,
                            TipoMovimiento.ASIGNACION,
                            origen=ubicacion_anterior,
                            destino=ubicacion_nueva,
                            responsable=personal,
                            observaciones="Importacion masiva desde Excel.",
                            request=request,
                        )
                        historial.registrar_historial(
                            request=request,
                            modulo=ModuloHistorial.ASIGNACION,
                            accion=historial.AccionHistorial.ASIGNACION,
                            titulo=f"Asignacion importada: {equipo} → {personal}",
                            objeto=asignacion,
                            entidad_relacionada=equipo,
                            enlace_nombre="equipo_detail",
                            enlace_pk=equipo.pk,
                        )
                        result["asignaciones_creadas"] += 1

                if row.tipo == TipoCategoriaInventario.EQUIPO:
                    result["equipos_creados"] += 1
                else:
                    result["perifericos_creados"] += 1
        except Exception as exc:
            result["errores_ejecucion"].append(
                f"Fila {row.row_num} ({row.codigo}): {exc}"
            )
            result["filas_omitidas"] += 1

    return result


def _style_header_row(ws, headers: list[str]):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = max(16, len(header) + 2)


def build_import_template() -> bytes:
    """Genera plantilla Excel con hojas de datos y catalogos de referencia."""
    wb = Workbook()

    ws_eq = wb.active
    ws_eq.title = SHEET_EQUIPOS
    _style_header_row(ws_eq, EQUIPOS_HEADERS)
    ws_eq.append(
        [
            "INV-001",
            "Laptop",
            "Dell",
            "Latitude 5540",
            "SN-ABC123",
            "En Stock",
            "",
            "",
            "",
            "Equipo de ejemplo",
            "Legado",
        ]
    )

    ws_per = wb.create_sheet(SHEET_PERIFERICOS)
    _style_header_row(ws_per, PERIFERICOS_HEADERS)
    ws_per.append(
        [
            "PER-001",
            "Monitor",
            "LG",
            "24MK430",
            "SN-MON001",
            "En Stock",
            "",
            "Monitor de ejemplo",
        ]
    )

    ws_cat = wb.create_sheet(SHEET_CATALOGOS)
    ws_cat.append(["Tipo", "Valor", "Detalle"])
    ws_cat["A1"].font = Font(bold=True)
    ws_cat["B1"].font = Font(bold=True)
    ws_cat["C1"].font = Font(bold=True)

    for cat in CategoriaEquipo.objects.filter(activo=True).order_by("tipo", "nombre_categoria"):
        ws_cat.append(["Categoria", cat.nombre_categoria, cat.tipo])

    for ub in Ubicacion.objects.filter(activo=True).select_related("edificio", "zona").order_by(
        "edificio__nombre_edificio", "zona__nombre_zona", "referencia"
    ):
        ws_cat.append(["Ubicacion", str(ub), ""])

    for pers in Personal.objects.filter(activo=True).order_by("numero_empleado")[:500]:
        nombre = " ".join(
            p for p in [pers.nombre, pers.apellido_paterno, pers.apellido_materno] if p
        )
        ws_cat.append(["Personal", pers.numero_empleado, nombre])

    for prov in Proveedor.objects.filter(activo=True).order_by("nombre_proveedor")[:200]:
        ws_cat.append(["Proveedor", prov.nombre_proveedor, prov.codigo_interno or ""])

    for value, label in EstadoEquipo.choices:
        ws_cat.append(["Estado", value, label])

    for value, label in OrigenAltaEquipo.choices:
        ws_cat.append(["Origen alta", value, label])

    ws_cat.column_dimensions["A"].width = 14
    ws_cat.column_dimensions["B"].width = 36
    ws_cat.column_dimensions["C"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
